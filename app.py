import os
import io
import csv
import json
import time
from datetime import datetime, date
from urllib.parse import urlparse

import pymysql

from flask import (
    Flask,
    jsonify,
    render_template,
    request,
    send_file,
    Response,
    stream_with_context
)

from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

from services.scanner import run_scan, scan_stock


# ============================================================
# BASE DIRECTORY / ENV
# ============================================================

BASE_DIR = os.path.dirname(
    os.path.abspath(__file__)
)

ENV_FILE = os.path.join(
    BASE_DIR,
    ".env"
)

load_dotenv(ENV_FILE)


# ============================================================
# DATABASE SETTINGS
# ============================================================

def load_db_settings():
    """
    Load MySQL connection from MYSQL_URL.

    Railway:
        MYSQL_URL=${{MySQL.MYSQL_URL}}

    Local .env:
        MYSQL_URL=mysql://user:password@127.0.0.1:3306/database

    MYSQL_URL is preferred because Railway automatically provides
    the complete connection string.
    """

    mysql_url = (
        os.getenv("MYSQL_URL")
        or os.getenv("MYSQLURL")
        or ""
    ).strip()

    if not mysql_url:

        raise RuntimeError(
            "MYSQL_URL is not configured.\n"
            "For Railway, create MYSQL_URL in the application "
            "service and reference your MySQL service's MYSQL_URL."
        )

    parsed = urlparse(mysql_url)

    host = parsed.hostname
    port = parsed.port or 3306
    user = parsed.username
    password = parsed.password
    database = (parsed.path or "").lstrip("/")

    if not host:

        raise RuntimeError(
            "MYSQL_URL does not contain a valid database host."
        )

    if not user:

        raise RuntimeError(
            "MYSQL_URL does not contain a database username."
        )

    if not database:

        raise RuntimeError(
            "MYSQL_URL does not contain a database name."
        )

    return {
        "url": mysql_url,
        "host": host,
        "port": port,
        "user": user,
        "password": password or "",
        "name": database
    }


def build_db_uri(settings):
    """
    Convert Railway MYSQL_URL into SQLAlchemy URL.
    """

    mysql_url = settings["url"]

    if mysql_url.startswith(
        "mysql+pymysql://"
    ):

        return mysql_url

    if mysql_url.startswith(
        "mysql://"
    ):

        return mysql_url.replace(
            "mysql://",
            "mysql+pymysql://",
            1
        )

    raise RuntimeError(
        "Unsupported MYSQL_URL format. "
        "Expected mysql:// or mysql+pymysql://"
    )


# ============================================================
# LOAD DATABASE SETTINGS
# ============================================================

_initial_db_settings = load_db_settings()

DB_NAME = _initial_db_settings["name"]
DB_USER = _initial_db_settings["user"]
DB_PASSWORD = _initial_db_settings["password"]
DB_HOST = _initial_db_settings["host"]
DB_PORT = _initial_db_settings["port"]


# ============================================================
# DATABASE DIAGNOSTIC
# ============================================================

print("")
print("=" * 60)
print("STOCK BREAKOUT 3")
print("=" * 60)

print(
    "ENV FILE       :",
    ENV_FILE
)

print(
    "ENV EXISTS     :",
    os.path.exists(ENV_FILE)
)

print(
    "MYSQL HOST     :",
    DB_HOST
)

print(
    "MYSQL USER     :",
    DB_USER
)

print(
    "MYSQL DATABASE :",
    DB_NAME
)

print(
    "MYSQL PORT     :",
    DB_PORT
)

print(
    "PASSWORD LOADED:",
    bool(DB_PASSWORD)
)

print("=" * 60)
print("")


# ============================================================
# FLASK APP
# ============================================================

app = Flask(__name__)


# ============================================================
# SQLALCHEMY
# ============================================================

app.config[
    "SQLALCHEMY_DATABASE_URI"
] = build_db_uri(
    _initial_db_settings
)

app.config[
    "SQLALCHEMY_TRACK_MODIFICATIONS"
] = False

# Connection pool settings
app.config[
    "SQLALCHEMY_ENGINE_OPTIONS"
] = {
    "pool_pre_ping": True,
    "pool_recycle": 280,
    "pool_timeout": 30,
    "pool_size": 5,
    "max_overflow": 10
}


db = SQLAlchemy(app)


# ============================================================
# STOCK MASTER TABLE
# ============================================================

class StockMaster(db.Model):

    __tablename__ = "stock_master"

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    date = db.Column(
        db.Date,
        nullable=False,
        index=True
    )

    symbol = db.Column(
        db.String(50),
        nullable=False,
        index=True
    )

    stock_name = db.Column(
        db.String(255),
        nullable=False
    )

    exchange = db.Column(
        db.String(10),
        nullable=False,
        default="NSE"
    )

    breakout_level = db.Column(
        db.Numeric(14, 2),
        nullable=False
    )

    stoploss = db.Column(
        db.Numeric(14, 2),
        nullable=False
    )

    current_price = db.Column(
        db.Numeric(14, 2),
        nullable=True
    )

    youtuber = db.Column(
        db.String(100),
        nullable=True
    )

    advisor = db.Column(
        db.String(100),
        nullable=True
    )

    category = db.Column(
        db.String(100),
        nullable=False,
        default="Breakouts",
        index=True
    )

    created_at = db.Column(
        db.DateTime,
        nullable=False,
        default=datetime.utcnow
    )

    updated_at = db.Column(
        db.DateTime,
        nullable=False,
        default=datetime.utcnow,
        onupdate=datetime.utcnow
    )

    __table_args__ = (

        db.UniqueConstraint(
            "symbol",
            "exchange",
            name="uq_symbol_exchange"
        ),

    )


# ============================================================
# DATABASE INITIALIZATION
# ============================================================

_db_initialized = False


def initialize_database():

    global _db_initialized

    if _db_initialized:

        return True

    try:

        with app.app_context():

            db.create_all()

        print("")
        print("=" * 60)
        print("MYSQL DATABASE READY")
        print("=" * 60)

        print(
            "MYSQL HOST     :",
            DB_HOST
        )

        print(
            "MYSQL USER     :",
            DB_USER
        )

        print(
            "MYSQL DATABASE :",
            DB_NAME
        )

        print(
            "MYSQL PORT     :",
            DB_PORT
        )

        print(
            "TABLE          :",
            "stock_master"
        )

        print("=" * 60)
        print("")

        _db_initialized = True

        return True

    except Exception as error:

        print("")
        print("=" * 60)
        print("MYSQL TABLE INITIALIZATION ERROR")
        print("=" * 60)
        print(error)
        print("=" * 60)
        print("")

        return False


@app.before_request
def _ensure_database_ready():

    if not _db_initialized:

        initialize_database()


# ============================================================
# HELPERS
# ============================================================

def clean_string(value):

    if value is None:

        return ""

    return str(value).strip()


def parse_date(value):

    if isinstance(
        value,
        datetime
    ):

        return value.date()

    if isinstance(
        value,
        date
    ):

        return value

    if value is None:

        raise ValueError(
            "Date is required"
        )

    value = str(
        value
    ).strip()

    if not value:

        raise ValueError(
            "Date is required"
        )

    formats = [

        "%d-%m-%Y",

        "%Y-%m-%d",

        "%d/%m/%Y",

        "%Y/%m/%d",

        "%d-%m-%y",

        "%d/%m/%y"

    ]

    for fmt in formats:

        try:

            return datetime.strptime(
                value,
                fmt
            ).date()

        except ValueError:

            continue

    raise ValueError(
        f"Invalid date: {value}. "
        f"Use DD-MM-YYYY."
    )


def parse_required_float(
    value,
    field_name,
    row_number=None
):

    if value is None:

        prefix = (
            f"Row {row_number}: "
            if row_number
            else ""
        )

        raise ValueError(
            f"{prefix}{field_name} is blank"
        )

    value = str(
        value
    ).strip()

    if not value:

        prefix = (
            f"Row {row_number}: "
            if row_number
            else ""
        )

        raise ValueError(
            f"{prefix}{field_name} is blank"
        )

    try:

        return float(value)

    except (
        ValueError,
        TypeError
    ):

        prefix = (
            f"Row {row_number}: "
            if row_number
            else ""
        )

        raise ValueError(
            f"{prefix}{field_name} "
            f"must be numeric. "
            f"Found: {value}"
        )


def parse_optional_float(value):

    if value is None:

        return None

    value = str(
        value
    ).strip()

    if not value:

        return None

    try:

        return float(value)

    except (
        ValueError,
        TypeError
    ):

        raise ValueError(
            f"Invalid numeric value: {value}"
        )


def stock_to_dict(
    stock,
    live_result=None
):

    live_result = (
        live_result
        or {}
    )

    current_price = (
        live_result.get(
            "current_price"
        )
    )

    if current_price is None:

        if stock.current_price is not None:

            current_price = float(
                stock.current_price
            )

    percent_diff = (
        live_result.get(
            "percent_diff"
        )
    )

    if (
        percent_diff is None
        and current_price is not None
        and stock.breakout_level is not None
    ):

        breakout = float(
            stock.breakout_level
        )

        if breakout != 0:

            percent_diff = round(

                (
                    (
                        float(current_price)
                        - breakout
                    )
                    / breakout
                ) * 100,

                2
            )

    status = live_result.get(
        "status"
    )

    if status is None:

        if current_price is None:

            status = "NOT SCANNED"

        elif (
            float(current_price)
            >= float(stock.breakout_level)
        ):

            status = "YES"

        else:

            status = "NO"

    return {

        "id":
            stock.id,

        "date":
            stock.date.strftime(
                "%d-%m-%Y"
            ),

        "symbol":
            stock.symbol,

        "stock_name":
            stock.stock_name,

        "exchange":
            stock.exchange,

        "breakout_level":
            float(
                stock.breakout_level
            ),

        "stoploss":
            float(
                stock.stoploss
            ),

        "current_price":
            current_price,

        "percent_diff":
            percent_diff,

        "youtuber":
            stock.youtuber or "",

        "advisor":
            stock.advisor or "",

        "category":
            stock.category,

        "status":
            status
    }


# ============================================================
# QUERY FILTER
# ============================================================

def get_filtered_query():

    category = clean_string(
        request.args.get(
            "category",
            ""
        )
    )

    search = clean_string(
        request.args.get(
            "search",
            ""
        )
    )

    query = StockMaster.query

    if category:

        query = query.filter(
            StockMaster.category
            == category
        )

    if search:

        pattern = (
            f"%{search}%"
        )

        query = query.filter(

            or_(

                StockMaster.symbol.ilike(
                    pattern
                ),

                StockMaster.stock_name.ilike(
                    pattern
                ),

                StockMaster.youtuber.ilike(
                    pattern
                ),

                StockMaster.advisor.ilike(
                    pattern
                ),

                StockMaster.category.ilike(
                    pattern
                )

            )
        )

    return query


# ============================================================
# HOME
# ============================================================

@app.get("/")
def index():

    return render_template(
        "index.html"
    )


# ============================================================
# CATEGORIES
# ============================================================

@app.get("/api/categories")
def get_categories():

    rows = (

        db.session
        .query(
            StockMaster.category
        )
        .filter(
            StockMaster.category.isnot(None)
        )
        .distinct()
        .order_by(
            StockMaster.category
        )
        .all()

    )

    categories = [

        row[0]

        for row in rows

        if row[0]
    ]

    return jsonify({

        "categories":
            categories
    })


# ============================================================
# GET STOCKS
# ============================================================

@app.get("/api/stocks")
def get_stocks():

    page = max(

        request.args.get(
            "page",
            1,
            type=int
        ),

        1
    )

    page_size = min(

        max(

            request.args.get(
                "page_size",
                100,
                type=int
            ),

            10

        ),

        500
    )

    query = (
        get_filtered_query()
        .order_by(

            StockMaster.date.desc(),

            StockMaster.symbol.asc()

        )
    )

    all_rows = query.all()

    total = len(
        all_rows
    )

    pages = max(

        (
            total
            + page_size
            - 1
        )
        // page_size,

        1
    )

    start = (
        page - 1
    ) * page_size

    rows = all_rows[
        start:
        start + page_size
    ]

    live_results = app.config.get(
        "LIVE_RESULTS",
        {}
    )

    items = [

        stock_to_dict(

            stock,

            live_results.get(
                stock.id
            )

        )

        for stock in rows

    ]

    return jsonify({

        "items":
            items,

        "page":
            page,

        "pages":
            pages,

        "total":
            total,

        "has_next":
            page < pages,

        "has_prev":
            page > 1
    })


# ============================================================
# STATS
# ============================================================

@app.get("/api/stats")
def get_stats():

    rows = (
        get_filtered_query()
        .all()
    )

    live_results = app.config.get(
        "LIVE_RESULTS",
        {}
    )

    items = [

        stock_to_dict(

            stock,

            live_results.get(
                stock.id
            )

        )

        for stock in rows

    ]

    breakout_items = [

        item

        for item in items

        if item["status"] == "YES"

        and item["percent_diff"]
        is not None

    ]

    top_gainer = None

    if breakout_items:

        top_gainer = max(

            breakout_items,

            key=lambda x:
                x["percent_diff"]

        )

    avg_gain = 0

    if breakout_items:

        avg_gain = round(

            sum(

                item["percent_diff"]

                for item
                in breakout_items

            )
            / len(
                breakout_items
            ),

            2
        )

    return jsonify({

        "master_stocks":
            len(rows),

        "stocks_scanned":
             len(live_results),

        "breakout_stocks":
            len(
                breakout_items
            ),

        "avg_gain":
            avg_gain,

        "top_gainer":

            (

                {

                    "symbol":
                        top_gainer[
                            "symbol"
                        ],

                    "percent_diff":
                        top_gainer[
                            "percent_diff"
                        ]

                }

                if top_gainer

                else None

            )

    })


# ============================================================
# LIVE RUN SCAN - SERVER SENT EVENTS
# ============================================================

@app.get("/api/run-scan-stream")
def run_breakout_scan_stream():

    category = clean_string(
        request.args.get(
            "category",
            ""
        )
    )

    query = StockMaster.query

    if category:

        query = query.filter(
            StockMaster.category
            == category
        )

    selected_stocks = (
        query
        .order_by(
            StockMaster.symbol.asc()
        )
        .all()
    )

    if not selected_stocks:

        return jsonify({

            "success":
                False,

            "message":
                "No stocks found for the selected category."

        }), 400

    @stream_with_context
    def generate():

        total = len(
            selected_stocks
        )

        results = {}

        scanned = 0
        breakouts = 0
        not_found = 0
        errors = 0

        app.config[
            "LIVE_RESULTS"
        ] = {}

        yield (
            "data: "
            + json.dumps({

                "type":
                    "start",

                "total":
                    total,

                "category":
                    category or "ALL"

            })
            + "\n\n"
        )

        for index, stock in enumerate(
            selected_stocks,
            start=1
        ):

            try:
                print(
                f"========== SCANNING {index}/{total}: "
                 f"{stock.symbol} ==========",
                flush=True
        )

                start_time = time.time()

                result = scan_stock(
                    stock
                )
                elapsed = time.time() - start_time   
                print( f"========== COMPLETED {index}/{total}: "f"{stock.symbol} ({elapsed:.2f}s) ==========",flush=True)         

                db.session.commit()

            except Exception as error:

                db.session.rollback()

                result = {

                    "status":
                        "ERROR",

                    "error":
                        str(error)

                }

            results[
                stock.id
            ] = result

            app.config[
                "LIVE_RESULTS"
            ] = results.copy()

            #if result.get(
             #   "current_price"
            #) is not None:

            scanned += 1

            if result.get(
                "status"
            ) == "YES":

                breakouts += 1

            elif result.get(
                "status"
            ) == "NOT FOUND":

                not_found += 1

            elif result.get(
                "status"
            ) == "ERROR":

                errors += 1

            yield (
                "data: "
                + json.dumps({

                    "type":
                        "stock",

                    "index":
                        index,

                    "total":
                        total,

                    "result":
                        result,

                    "summary": {

                        "scanned":
                            scanned,

                        "breakouts":
                            breakouts,

                        "not_found":
                            not_found,

                        "errors":
                            errors
                    }

                })
                + "\n\n"
            )

       

        yield (
            "data: "
            + json.dumps({

                "type":
                    "complete",

                "total":
                    total,

                "summary": {

                    "scanned":
                        scanned,

                    "breakouts":
                        breakouts,

                    "not_found":
                        not_found,

                    "errors":
                        errors
                }

            })
            + "\n\n"
        )

    return Response(

        generate(),

        mimetype="text/event-stream",

        headers={

            "Cache-Control":
                "no-cache",

            "X-Accel-Buffering":
                "no",

            "Connection":
                "keep-alive"

        }

    )


# ============================================================
# RUN SCAN
# ============================================================

@app.post("/api/run-scan")
def run_breakout_scan():

    category = clean_string(

        request.args.get(
            "category",
            ""
        )

    )

    query = StockMaster.query

    if category:

        query = query.filter(

            StockMaster.category
            == category

        )

    selected_stocks = (

        query
        .order_by(
            StockMaster.symbol.asc()
        )
        .all()

    )

    if not selected_stocks:

        return jsonify({

            "success":
                False,

            "message":
                "No stocks found for the selected category.",

            "stocks_scanned":
                0,

            "breakouts_found":
                0,

            "not_found":
                0,

            "errors":
                0

        }), 400

    print("")
    print("=" * 60)
    print("FLASK SCAN REQUEST")
    print("=" * 60)

    print(
        "Category:",
        category or "ALL"
    )

    print(
        "Stocks selected:",
        len(selected_stocks)
    )

    print("=" * 60)

    try:

        result = run_scan(
            selected_stocks
        )

    except Exception as error:

        return jsonify({

            "success":
                False,

            "message":
                str(error)

        }), 500

    app.config[
        "LIVE_RESULTS"
    ] = result.get(
        "results",
        {}
    )

    return jsonify({

        "success":
            True,

        "category":
            category or "ALL",

        "stocks_selected":
            len(
                selected_stocks
            ),

        "stocks_scanned":
            result.get(
                "stocks_scanned",
                0
            ),

        "breakouts_found":
            result.get(
                "breakouts_found",
                0
            ),

        "not_found":
            result.get(
                "not_found",
                0
            ),

        "errors":
            result.get(
                "errors",
                0
            ),

        "results":
            result.get(
                "results",
                {}
            )

    })


# ============================================================
# ADD STOCK
# ============================================================

@app.post("/api/stocks")
def create_stock():

    data = (
        request.get_json()
        or {}
    )

    try:

        symbol = clean_string(
            data.get(
                "symbol"
            )
        ).upper()

        if not symbol:

            raise ValueError(
                "Symbol is required"
            )

        exchange = clean_string(

            data.get(
                "exchange"
            )

            or "NSE"

        ).upper()

        existing = (

            StockMaster.query

            .filter_by(

                symbol=symbol,

                exchange=exchange

            )

            .first()

        )

        if existing:

            return jsonify({

                "success":
                    False,

                "message":
                    (
                        f"{symbol} already exists "
                        f"in {exchange}."
                    )

            }), 400

        stock = StockMaster(

            date=parse_date(
                data.get(
                    "date"
                )
            ),

            symbol=symbol,

            stock_name=clean_string(
                data.get(
                    "stock_name"
                )
            ),

            exchange=exchange,

            breakout_level=
                parse_required_float(

                    data.get(
                        "breakout_level"
                    ),

                    "Breakout Level"

                ),

            stoploss=
                parse_required_float(

                    data.get(
                        "stoploss"
                    ),

                    "StopLoss"

                ),

            current_price=
                parse_optional_float(

                    data.get(
                        "current_price"
                    )

                ),

            youtuber=clean_string(

                data.get(
                    "youtuber"
                )

            ),

            advisor=clean_string(

                data.get(
                    "advisor"
                )

            ),

            category=(

                clean_string(

                    data.get(
                        "category"
                    )

                )

                or "Breakouts"

            )

        )

        db.session.add(
            stock
        )

        db.session.commit()

        return jsonify({

            "success":
                True,

            "item":
                stock_to_dict(
                    stock
                )

        }), 201

    except Exception as error:

        db.session.rollback()

        return jsonify({

            "success":
                False,

            "message":
                str(error)

        }), 400


# ============================================================
# UPDATE STOCK
# ============================================================

@app.put(
    "/api/stocks/<int:stock_id>"
)
def update_stock(stock_id):

    stock = (
        db.session.get(
            StockMaster,
            stock_id
        )
    )

    if stock is None:

        return jsonify({

            "success":
                False,

            "message":
                "Stock not found."

        }), 404

    data = (
        request.get_json()
        or {}
    )

    try:

        symbol = clean_string(

            data.get(
                "symbol"
            )

        ).upper()

        if not symbol:

            raise ValueError(
                "Symbol is required"
            )

        exchange = clean_string(

            data.get(
                "exchange"
            )

            or "NSE"

        ).upper()

        duplicate = (

            StockMaster.query

            .filter(

                StockMaster.symbol
                == symbol,

                StockMaster.exchange
                == exchange,

                StockMaster.id
                != stock.id

            )

            .first()

        )

        if duplicate:

            return jsonify({

                "success":
                    False,

                "message":
                    (
                        f"{symbol} already exists "
                        f"in {exchange}."
                    )

            }), 400

        stock.date = parse_date(

            data.get(
                "date"
            )

        )

        stock.symbol = symbol

        stock.stock_name = clean_string(

            data.get(
                "stock_name"
            )

        )

        stock.exchange = exchange

        stock.breakout_level = (

            parse_required_float(

                data.get(
                    "breakout_level"
                ),

                "Breakout Level"

            )

        )

        stock.stoploss = (

            parse_required_float(

                data.get(
                    "stoploss"
                ),

                "StopLoss"

            )

        )

        stock.current_price = (

            parse_optional_float(

                data.get(
                    "current_price"
                )

            )

        )

        stock.youtuber = clean_string(

            data.get(
                "youtuber"
            )

        )

        stock.advisor = clean_string(

            data.get(
                "advisor"
            )

        )

        stock.category = (

            clean_string(

                data.get(
                    "category"
                )

            )

            or "Breakouts"

        )

        stock.updated_at = (
            datetime.utcnow()
        )

        db.session.commit()

        return jsonify({

            "success":
                True,

            "item":
                stock_to_dict(
                    stock
                )

        })

    except Exception as error:

        db.session.rollback()

        return jsonify({

            "success":
                False,

            "message":
                str(error)

        }), 400


# ============================================================
# DELETE STOCK
# ============================================================

@app.delete(
    "/api/stocks/<int:stock_id>"
)
def delete_stock(stock_id):

    stock = (
        db.session.get(
            StockMaster,
            stock_id
        )
    )

    if stock is None:

        return jsonify({

            "success":
                False,

            "message":
                "Stock not found."

        }), 404

    try:

        db.session.delete(
            stock
        )

        db.session.commit()

        live_results = app.config.get(
            "LIVE_RESULTS",
            {}
        )

        live_results.pop(
            stock_id,
            None
        )

        app.config[
            "LIVE_RESULTS"
        ] = live_results

        return jsonify({

            "success":
                True

        })

    except Exception as error:

        db.session.rollback()

        return jsonify({

            "success":
                False,

            "message":
                str(error)

        }), 500


# ============================================================
# UPLOAD EXCEL / CSV
# ============================================================

@app.post("/api/upload")
def upload_file():

    uploaded_file = (
        request.files.get(
            "file"
        )
    )

    if not uploaded_file:

        return jsonify({

            "success":
                False,

            "message":
                "Please select a file."

        }), 400

    filename = (
        uploaded_file.filename
        or ""
    ).lower()

    if not (

        filename.endswith(
            ".xlsx"
        )

        or

        filename.endswith(
            ".csv"
        )

    ):

        return jsonify({

            "success":
                False,

            "message":
                "Only XLSX and CSV files are supported."

        }), 400

    try:

        rows = []

        # ====================================================
        # EXCEL
        # ====================================================

        if filename.endswith(
            ".xlsx"
        ):

            workbook = load_workbook(

                uploaded_file,

                read_only=True,

                data_only=True

            )

            worksheet = (
                workbook.active
            )

            values = list(
                worksheet.values
            )

            if not values:

                return jsonify({

                    "success":
                        False,

                    "message":
                        "Excel file is empty."

                }), 400

            headers = []

            for value in values[0]:

                header = clean_string(
                    value
                )

                header = (
                    header
                    .lower()
                    .replace(
                        "\n",
                        " "
                    )
                    .strip()
                )

                headers.append(
                    header
                )

            for row_number, row_values in enumerate(

                values[1:],

                start=2

            ):

                row = dict(

                    zip(
                        headers,
                        row_values
                    )

                )

                rows.append(

                    (
                        row_number,
                        row
                    )

                )

        # ====================================================
        # CSV
        # ====================================================

        else:

            text_file = (
                io.TextIOWrapper(

                    uploaded_file.stream,

                    encoding="utf-8-sig"

                )
            )

            reader = csv.DictReader(
                text_file
            )

            for row_number, row in enumerate(

                reader,

                start=2

            ):

                cleaned = {}

                for key, value in row.items():

                    cleaned[
                        clean_string(
                            key
                        )
                        .lower()
                        .strip()
                    ] = value

                rows.append(

                    (
                        row_number,
                        cleaned
                    )

                )

        # ====================================================
        # COLUMN HELPER
        # ====================================================

        def get_value(
            row,
            *names
        ):

            for name in names:

                normalized = (

                    clean_string(
                        name
                    )
                    .lower()
                    .replace(
                        "\n",
                        " "
                    )
                    .strip()

                )

                if normalized in row:

                    return row[
                        normalized
                    ]

            return None

        # ====================================================
        # COUNTERS
        # ====================================================

        added = 0
        updated = 0
        skipped = 0

        errors = []

        # ====================================================
        # PROCESS ROWS
        # ====================================================

        for row_number, row in rows:

            try:

                symbol = clean_string(

                    get_value(
                        row,
                        "symbol"
                    )

                ).upper()

                if not symbol:

                    continue

                exchange = clean_string(

                    get_value(
                        row,
                        "exchange"
                    )

                    or "NSE"

                ).upper()

                stock_date = parse_date(

                    get_value(
                        row,
                        "date"
                    )

                )

                stock_name = clean_string(

                    get_value(

                        row,

                        "stocks",

                        "stock",

                        "stock name",

                        "stock_name"

                    )

                )

                if not stock_name:

                    raise ValueError(
                        "Stocks/Stock Name is blank"
                    )

                breakout_level = (

                    parse_required_float(

                        get_value(

                            row,

                            "breakout level",

                            "breakout_level",

                            "breakout"

                        ),

                        "BreakOut Level",

                        row_number

                    )

                )

                stoploss = (

                    parse_required_float(

                        get_value(

                            row,

                            "stoploss",

                            "stop loss",

                            "stop_loss"

                        ),

                        "StopLoss",

                        row_number

                    )

                )

                current_price_value = (

                    get_value(

                        row,

                        "current price",

                        "current_price",

                        "currentprice"

                    )

                )

                current_price = (

                    parse_optional_float(

                        current_price_value

                    )

                )

                youtuber = clean_string(

                    get_value(

                        row,

                        "you tuber",

                        "youtuber",

                        "youtube",

                        "you_tuber"

                    )

                )

                advisor = clean_string(

                    get_value(

                        row,

                        "advisor"

                    )

                )

                category = clean_string(

                    get_value(

                        row,

                        "category"

                    )

                )

                if not category:

                    category = "Breakouts"

                stock = (

                    StockMaster.query

                    .filter_by(

                        symbol=symbol,

                        exchange=exchange

                    )

                    .first()

                )

                # =================================================
                # INSERT
                # =================================================

                if stock is None:

                    stock = StockMaster(

                        date=stock_date,

                        symbol=symbol,

                        stock_name=stock_name,

                        exchange=exchange,

                        breakout_level=
                            breakout_level,

                        stoploss=
                            stoploss,

                        current_price=
                            current_price,

                        youtuber=
                            youtuber,

                        advisor=
                            advisor,

                        category=
                            category

                    )

                    db.session.add(
                        stock
                    )

                    added += 1

                # =================================================
                # UPDATE
                # =================================================

                else:

                    stock.date = (
                        stock_date
                    )

                    stock.stock_name = (
                        stock_name
                    )

                    stock.exchange = (
                        exchange
                    )

                    stock.breakout_level = (
                        breakout_level
                    )

                    stock.stoploss = (
                        stoploss
                    )

                    # Preserve existing current price
                    # when Excel doesn't contain one.

                    if current_price is not None:

                        stock.current_price = (
                            current_price
                        )

                    stock.youtuber = (
                        youtuber
                    )

                    stock.advisor = (
                        advisor
                    )

                    stock.category = (
                        category
                    )

                    stock.updated_at = (
                        datetime.utcnow()
                    )

                    updated += 1

            except Exception as row_error:

                skipped += 1

                errors.append({

                    "row":
                        row_number,

                    "symbol":
                        clean_string(

                            get_value(
                                row,
                                "symbol"
                            )

                        ),

                    "error":
                        str(
                            row_error
                        )

                })

        db.session.commit()

        app.config[
            "LIVE_RESULTS"
        ] = {}

        return jsonify({

            "success":
                True,

            "added":
                added,

            "updated":
                updated,

            "skipped":
                skipped,

            "errors":
                errors,

            "message":
                (
                    f"Upload completed. "
                    f"{added} added, "
                    f"{updated} updated, "
                    f"{skipped} skipped."
                )

        })

    except Exception as error:

        db.session.rollback()

        return jsonify({

            "success":
                False,

            "message":
                f"Upload failed: {error}"

        }), 400


# ============================================================
# DOWNLOAD EXCEL
# ============================================================

@app.get("/api/download-excel")
def download_excel():

    rows = (
        get_filtered_query()
        .order_by(
            StockMaster.date.desc(),
            StockMaster.symbol.asc()
        )
        .all()
    )

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Breakout Stocks"

    worksheet.append([

        "Date",

        "Symbol",

        "Stocks",

        "Exchange",

        "BreakOut Level",

        "StopLoss",

        "Current Price",

        "You Tuber",

        "Advisor",

        "Category"

    ])

    for stock in rows:

        if stock.current_price is None:

            continue

        if stock.breakout_level is None:

            continue

        try:

            current_price = float(
                stock.current_price
            )

            breakout_level = float(
                stock.breakout_level
            )

        except (
            TypeError,
            ValueError
        ):

            continue

        if breakout_level <= 0:

            continue

        if current_price < breakout_level:

            continue

        worksheet.append([

            stock.date,

            stock.symbol,

            stock.stock_name,

            stock.exchange,

            breakout_level,

            (
                float(stock.stoploss)
                if stock.stoploss is not None
                else None
            ),

            current_price,

            stock.youtuber,

            stock.advisor,

            stock.category

        ])

    output = io.BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return send_file(

        output,

        as_attachment=True,

        download_name="breakout_stocks.xlsx",

        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )

    )


# ============================================================
# HEALTH CHECK
# ============================================================

@app.get("/api/health")
def health():

    try:

        db.session.execute(
            db.text(
                "SELECT 1"
            )
        )

        stock_count = (
            StockMaster.query.count()
        )

        return jsonify({

            "success":
                True,

            "database":
                "connected",

            "stock_count":
                stock_count,

            "host":
                DB_HOST,

            "database_name":
                DB_NAME

        })

    except Exception as error:

        return jsonify({

            "success":
                False,

            "database":
                "error",

            "message":
                str(error),

            "host":
                DB_HOST

        }), 500


# ============================================================
# APPLICATION START
# ============================================================

if __name__ == "__main__":

    app.run(

        host="0.0.0.0",

        port=int(

            os.getenv(
                "PORT",
                "5000"
            )

        ),

        debug=False

    )

